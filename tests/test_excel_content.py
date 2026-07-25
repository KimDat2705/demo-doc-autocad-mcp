# -*- coding: utf-8 -*-
"""KHOÁ NỘI DUNG EXCEL (xuat_excel) — TẤT ĐỊNH, KHÔNG gọi Gemini/mạng, KHÔNG tốn phí.
Chạy:  python tests/test_excel_content.py

Trước đây chỉ có smoke 'xuat_excel KHÔNG crash + có file_id' (nhóm D). File này MỞ LẠI .xlsx bằng
openpyxl và khoá NỘI DUNG THẬT:
  A. STRUCTURE — header đúng 8 cột (gồm 'Chưa chắc'); có khối 'TỔNG PHỤ'; có ≥1 dòng dữ liệu.
  B. LEARNED KHÔNG VÀO BÀN GIAO — dạy 1 quy ước học (P3 KG_PER_UNIT) trên residual 'S1 (1 bo) 12.5 kg':
     số 12.5 CHỈ nằm ở khối 'CHƯA XÁC NHẬN', KHÔNG ở dòng 'TỔNG' (tổng phụ), handle học KHÔNG ở bảng chính 8 cột.
(Bất biến P4: số HỌC hiện để đối tác thấy nhưng TUYỆT ĐỐI không cộng vào tổng/Excel bàn giao.)"""
import os, sys, io
os.environ.setdefault("READFILE_MAX_MB", "300")
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.normpath(os.path.join(HERE, "..")))
from tools_core import Drawing, RENDER_DIR

BASE = os.path.normpath(os.path.join(HERE, "..", "..", "input_files", "_dxf"))
# Ten thu muc/file that giu NGOAI repo (gitignored) — xem corpus_local.example.py
try:
    from corpus_local import KT, KC
except Exception:
    KT = KC = ""

PASS = FAIL = SKIP = 0


def _emit(name, ok, extra=""):
    global PASS, FAIL
    okk = bool(ok); PASS += int(okk); FAIL += int(not okk)
    print("  [%s] %s%s" % ("OK" if okk else "FAIL", name, (" " + extra) if extra else ""))


def skip(nhom, ly_do):
    global SKIP
    SKIP += 1
    print("  [..] BO QUA %s — %s (KHONG phai da kiem)" % (nhom, ly_do))


def _rows_of(file_id):
    """Mở .xlsx vừa xuất -> list[tuple] các dòng (values_only) của sheet ACTIVE. None nếu thiếu file."""
    from openpyxl import load_workbook
    fp = os.path.join(RENDER_DIR, file_id)
    if not os.path.isfile(fp):
        return None
    wb = load_workbook(fp, read_only=True, data_only=True)
    ws = wb.active
    rows = [tuple(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    return rows


def _rows_of_sheet(file_id, sheet_name):
    """Đọc 1 sheet CỤ THỂ theo tên (dùng cho sheet 'Tien_luong' I2). None nếu thiếu file/sheet."""
    from openpyxl import load_workbook
    fp = os.path.join(RENDER_DIR, file_id)
    if not os.path.isfile(fp):
        return None
    wb = load_workbook(fp, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close(); return None
    rows = [tuple(r) for r in wb[sheet_name].iter_rows(values_only=True)]
    wb.close()
    return rows


def _is_empty(row):
    return all(c is None or (isinstance(c, str) and c.strip() == "") for c in row)


def _cells_str(row):
    return [("" if c is None else str(c)) for c in row]


_MADE = []   # file_id đã xuất (dọn cuối)


def main():
    global PASS, FAIL, SKIP
    if not (os.path.isfile(KT) and os.path.isfile(KC)):
        skip("ALL", "khong thay fixture KT/KC (%s)" % BASE)
        print("\n%d PASS / %d FAIL / %d BO QUA" % (PASS, FAIL, SKIP))
        return 1 if FAIL else 0
    # openpyxl bắt buộc để đọc lại nội dung — thiếu thì SKIP (không crash)
    try:
        import openpyxl  # noqa: F401
    except Exception:
        skip("ALL", "may chua cai openpyxl -> khong doc lai duoc .xlsx")
        print("\n%d PASS / %d FAIL / %d BO QUA" % (PASS, FAIL, SKIP))
        return 1 if FAIL else 0

    kt, kc = Drawing(KT), Drawing(KC)

    print("[A] STRUCTURE — header 8 cột + TỔNG PHỤ + ≥1 dòng dữ liệu (KT & KC)")
    for ten, dwg in [("KT", kt), ("KC", kc)]:
        xl = dwg.xuat_excel()
        fid = xl.get("file_id")
        if not fid:
            _emit("%s: xuat_excel trả file_id" % ten, False, "-> loi=%s" % xl.get("loi"))
            continue
        _MADE.append(fid)
        rows = _rows_of(fid)
        if rows is None:
            _emit("%s: file .xlsx tồn tại trên đĩa" % ten, False)
            continue
        header = _cells_str(rows[0])
        _emit("%s: header đúng 8 cột" % ten, len(rows[0]) == 8, "-> %d cột" % len(rows[0]))
        _emit("%s: header khớp thứ tự + có 'Chưa chắc' + 'Handle'" % ten,
              header[:8] == ["STT", "Hạng mục", "Loại", "Giá trị", "Đơn vị", "Nguồn", "Chưa chắc", "Handle"])
        # dòng dữ liệu bảng chính = từ sau header tới dòng rỗng đầu tiên
        body = []
        for r in rows[1:]:
            if _is_empty(r):
                break
            body.append(r)
        _emit("%s: có ≥1 dòng dữ liệu bảng chính (STT là số)" % ten,
              len(body) >= 1 and isinstance(body[0][0], (int, float)), "-> %d dòng" % len(body))
        _emit("%s: có khối 'TỔNG PHỤ'" % ten,
              any("TỔNG PHỤ" in c for r in rows for c in _cells_str(r)))
        _emit("%s: có ≥1 dòng tổng phụ 'TỔNG <loại>' (hệ cộng theo loại+đơn vị)" % ten,
              any(_cells_str(r)[1].startswith("TỔNG ") for r in rows if len(r) > 1 and r[1]))

    print("[B] LEARNED (P3) KHÔNG VÀO BÀN GIAO — 12.5 chỉ ở khối 'CHƯA XÁC NHẬN', handle học vắng bảng chính")
    _HH = "ZZEXCEL1"
    # residual XA (x,y lớn -> không near mã/không index) + nhãn per-unit 'S1 (1 bo) 12.5 kg'
    kt.texts.append({"handle": _HH, "vn": "S1 (1 bo) 12.5 kg", "x": 8.0e8, "y": 8.0e8})
    kt._text_by_handle[_HH] = kt.texts[-1]
    kt.used_handles = kt._build_used_handles()
    rhk = kt.hoc_quy_uoc(_HH, "KG_PER_UNIT", "S1")
    _emit("B0: hoc_quy_uoc hợp lệ (residual S1 12.5 kg) -> ok + preview 12.5",
          rhk.get("ok") and abs((rhk.get("ung_vien_xem_truoc") or {}).get("gia_tri", 0) - 12.5) < 0.01,
          "-> %s" % rhk.get("ung_vien_xem_truoc"))

    xl = kt.xuat_excel()
    fid = xl.get("file_id")
    if not fid:
        _emit("B: xuat_excel (sau khi học) trả file_id", False, "-> loi=%s" % xl.get("loi"))
    else:
        _MADE.append(fid)
        rows = _rows_of(fid)
        if rows is None:
            _emit("B: file .xlsx tồn tại trên đĩa", False)
        else:
            # bảng chính = header..dòng rỗng đầu tiên
            body = []
            for r in rows[1:]:
                if _is_empty(r):
                    break
                body.append(r)
            # vị trí header khối 'CHƯA XÁC NHẬN'
            blk_start = None
            for i, r in enumerate(rows):
                if any("CHƯA XÁC NHẬN" in c and "QUY ƯỚC" in c for c in _cells_str(r)):
                    blk_start = i
                    break
            _emit("B1: có khối header 'QUY ƯỚC ĐỐI TÁC DẠY (CHƯA XÁC NHẬN…)'", blk_start is not None)
            # dòng 12.5 nằm trong khối chưa-xác-nhận, cột Handle = _HH, y_nghia kg_moi_bo
            blk_rows = rows[blk_start + 1:] if blk_start is not None else []
            hit = None
            for r in blk_rows:
                if _is_empty(r):
                    break
                if len(r) >= 8 and isinstance(r[3], (int, float)) and abs(r[3] - 12.5) < 0.01:
                    hit = r
                    break
            _emit("B2: 12.5 nằm ở khối CHƯA XÁC NHẬN + Handle=%s + có 'kg_moi_bo'" % _HH,
                  hit is not None and str(hit[7]) == _HH and "kg_moi_bo" in _cells_str(hit)[1],
                  "-> %s" % (None if hit is None else _cells_str(hit)))
            # 12.5 KHÔNG lọt vào bất kỳ dòng TỔNG (tổng phụ)
            tong_rows = [r for r in rows if len(r) > 1 and r[1] and str(r[1]).startswith("TỔNG ")]
            _emit("B3: 12.5 KHÔNG nằm ở dòng 'TỔNG …' (không cộng số học vào tổng)",
                  not any(isinstance(c, (int, float)) and abs(c - 12.5) < 0.01
                          for r in tong_rows for c in r if not isinstance(c, bool)))
            # handle học KHÔNG xuất hiện ở cột Handle của BẢNG CHÍNH 8 cột
            _emit("B4: handle học %s KHÔNG ở cột Handle bảng chính (không vào bàn giao)" % _HH,
                  all(_cells_str(r)[7] != _HH for r in body))
            # 12.5 xuất hiện đúng 1 lần trên toàn sheet (chỉ khối chưa-xác-nhận), không rò chỗ khác
            n125 = sum(1 for r in rows for c in r
                       if isinstance(c, (int, float)) and not isinstance(c, bool) and abs(c - 12.5) < 0.01)
            _emit("B5: 12.5 xuất hiện đúng 1 lần trên sheet (chỉ khối chưa-xác-nhận)", n125 == 1, "-> %d lần" % n125)

    # ── [C] I2: sheet "Tien_luong" (BOQ phẳng copy-ready) — THÊM SHEET không cướp active, KHÔNG đơn giá, P4 số-học KHÔNG lọt,
    # subtotal KHỚP bảng chính (cùng nguồn số). `fid` ở đây là bản xuất KHI ĐANG DẠY 12.5 (trước thu_hoi) -> test P4 mạnh nhất.
    print("[C] I2 SHEET 'Tien_luong' — thêm sheet an toàn (không cướp active) + không đơn giá + P4 + subtotal khớp")
    if fid:
        from openpyxl import load_workbook
        _wb = load_workbook(os.path.join(RENDER_DIR, fid), read_only=True, data_only=True)
        _sheets, _active = _wb.sheetnames, _wb.active.title
        _wb.close()
        _emit("C1: có sheet 'Tien_luong' + wb.active VẪN 'Tong hop khoi luong' (không cướp active -> [A] an toàn)",
              "Tien_luong" in _sheets and _active == "Tong hop khoi luong", "-> sheets=%s active=%s" % (_sheets, _active))
        tl = _rows_of_sheet(fid, "Tien_luong") or []
        _hdr = list(tl[0]) if tl else []
        _emit("C2: header Tien_luong KHÔNG có cột Đơn giá/Thành tiền (phạm vi: CHỈ khối lượng)",
              bool(_hdr) and not any(("đơn giá" in str(c).lower() or "thành tiền" in str(c).lower()) for c in _hdr),
              "-> %s" % _hdr)
        _n125_tl = sum(1 for r in tl for c in r
                       if isinstance(c, (int, float)) and not isinstance(c, bool) and abs(c - 12.5) < 0.01)
        _emit("C3 (P4): số học 12.5 (ĐANG DẠY) KHÔNG lọt vào Tien_luong (chỉ dùng th['bang'], không quy_uoc)",
              _n125_tl == 0, "-> %d lần" % _n125_tl)
        _main = _rows_of(fid) or []
        _main_vals = sorted(round(float(r[3]), 3) for r in _main
                            if len(r) > 3 and r[1] and str(r[1]).startswith("TỔNG ")
                            and isinstance(r[3], (int, float)) and not isinstance(r[3], bool))
        _tl_vals = sorted(round(float(r[4]), 3) for r in tl
                          if len(r) > 4 and r[2] and str(r[2]).startswith("Cộng ")
                          and isinstance(r[4], (int, float)) and not isinstance(r[4], bool))
        _emit("C4: subtotal Tien_luong ('Cộng …') KHỚP TỔNG PHỤ bảng chính (cùng nguồn số, KHÔNG lệch 2 sheet)",
              _tl_vals == _main_vals and len(_tl_vals) > 0, "-> tl=%s main=%s" % (_tl_vals, _main_vals))
    else:
        skip("C", "khong co fid de kiem sheet Tien_luong")

    kt.thu_hoi_quy_uoc("")
    # sau thu hồi: xuất lại -> KHÔNG còn khối chưa-xác-nhận, không còn 12.5
    xl = kt.xuat_excel()
    fid = xl.get("file_id")
    if fid:
        _MADE.append(fid)
        rows = _rows_of(fid) or []
        _emit("B6: sau thu_hoi_quy_uoc -> KHÔNG còn khối CHƯA XÁC NHẬN + KHÔNG còn 12.5",
              (not any("CHƯA XÁC NHẬN" in c for r in rows for c in _cells_str(r)))
              and (not any(isinstance(c, (int, float)) and not isinstance(c, bool) and abs(c - 12.5) < 0.01
                           for r in rows for c in r)))
    else:
        _emit("B6: xuat_excel sau thu hồi trả file_id", False)

    # DỌN file .xlsx test tạo ra
    for f in _MADE:
        try:
            os.remove(os.path.join(RENDER_DIR, f))
        except OSError:
            pass

    if SKIP:
        print("CANH BAO: %d nhom BO QUA do thieu fixture — KHONG phai da kiem" % SKIP)
    print("\n%d PASS / %d FAIL / %d BO QUA" % (PASS, FAIL, SKIP))
    return 1 if FAIL else 0


if __name__ == "__main__":
    sys.exit(main())
