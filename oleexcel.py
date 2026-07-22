# -*- coding: utf-8 -*-
"""Đọc bảng Excel NHÚNG (OLE2FRAME) trong bản vẽ CAD — đường CHÍNH = binary (chính xác tuyệt đối,
KHÔNG OCR). Vá điểm yếu: 19/65 file corpus dán bảng thép/khối-lượng dạng OLE mà ezdxf chỉ thấy là khối
nhúng → engine đọc 0. Probe corpus thật (2026-07-22): 89% OLE là Excel đọc được bằng đường này.

Kiến trúc (khớp cây quyết định U3 B0–B6 trong PHUONG_AN_NANG_CAP_DU_AN.md):
  binary_data() → tìm magic CFBF (có header thừa, KHÔNG ở offset 0) → olefile →
    Workbook/Book → xls (xlrd, decode KHOAN DUNG) ·  Package/CONTENTS → tìm .xlsx/.xls lồng bên trong
  Không phải Excel (ảnh StaticDib/PBrush/EMF, packager lạ) → loai='anh'/'khac' để CALLER OCR/cảnh báo.

Số ĐỌC RA chính xác (đọc nguyên bytes); CHỮ có thể garble TCVN3 → caller áp vntext.to_unicode.
Thà LỘ 'không đọc được' còn hơn phát số sai: mọi nhánh fail-soft, không ném ra ngoài."""
import io

# --- xlrd: decode KHOAN DUNG (errors='replace') — vài .xls VN làm xlrd 2.x chết ở utf-16 surrogate;
#     chữ hỏng thành U+FFFD (caller lọc), SỐ vẫn nguyên. Vá tại import, 1 lần. ---
try:
    import xlrd
    import xlrd.biffh
    import xlrd.book
    import xlrd.compdoc
    _lenient_unicode = lambda b, enc: b.decode(enc, "replace")
    for _m in (xlrd.biffh, xlrd.book, xlrd.compdoc):
        _m.unicode = _lenient_unicode
except Exception:  # pragma: no cover
    xlrd = None
try:
    import olefile
except Exception:  # pragma: no cover
    olefile = None
try:
    import openpyxl
except Exception:  # pragma: no cover
    openpyxl = None

CFBF = bytes.fromhex("D0CF11E0A1B11AE1")   # OLE2 Compound File magic
PKZIP = b"PK\x03\x04"                       # .xlsx (zip) magic
_MAX_ROWS = 2000                            # trần chống bảng khổng lồ ăn RAM
_MAX_COLS = 60


def _rows_from_xls(data):
    """BIFF .xls (bytes stream Workbook) -> list[list]. None nếu không đọc được."""
    if xlrd is None:
        return None
    try:
        bk = xlrd.open_workbook(file_contents=data)
    except Exception:
        return None
    out = []
    sh = bk.sheet_by_index(0)
    for r in range(min(sh.nrows, _MAX_ROWS)):
        out.append([sh.cell_value(r, c) for c in range(min(sh.ncols, _MAX_COLS))])
    return {"sheet": sh.name, "nrows": sh.nrows, "ncols": sh.ncols, "rows": out}


def _rows_from_xlsx(data):
    """.xlsx (bytes zip) -> list[list]. None nếu không đọc được."""
    if openpyxl is None:
        return None
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception:
        return None
    ws = wb.worksheets[0]
    out = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= _MAX_ROWS:
            break
        out.append(["" if v is None else v for v in row[:_MAX_COLS]])
    res = {"sheet": ws.title, "nrows": ws.max_row or len(out), "ncols": ws.max_column or 0, "rows": out}
    wb.close()
    return res


def _progid(ole):
    """ProgID (app nguồn) từ \x01CompObj — chỉ để phân loại/log, KHÔNG dùng làm số."""
    try:
        if ole.exists("\x01CompObj"):
            raw = ole.openstream("\x01CompObj").read().decode("latin-1", "replace")
            import re
            for c in re.findall(r"[ -~]{4,}", raw):
                if any(k in c for k in (".Sheet", ".Document", "Package", "Excel", "Word", "Acro", "Picture", "Bitmap", "Dib", "Brush")):
                    return c.strip()
    except Exception:
        pass
    return ""


def doc_tu_blob(blob):
    """Blob OLE thô -> dict {loai, ...}. loai: 'excel' (rows) | 'anh' | 'khac' | 'rong'.
    KHÔNG ném lỗi. 'excel' kèm sheet/nrows/ncols/rows (SỐ chính xác, chữ có thể garble)."""
    if not blob:
        return {"loai": "rong"}
    # (1) .xlsx zip trực tiếp
    if blob[:4] == PKZIP:
        t = _rows_from_xlsx(blob)
        return {"loai": "excel", "nguon_stream": "zip", **t} if t else {"loai": "khac", "ghi_chu": "PKzip khong doc duoc"}
    off = blob.find(CFBF)
    if off < 0:
        # .xlsx zip LỒNG ở offset bất kỳ (vd gói packager không bọc CFBF) — openpyxl XÁC THỰC zip
        # nên không false-positive từ vài byte 'PK' ngẫu nhiên.
        z = blob.find(PKZIP)
        if z >= 0:
            t = _rows_from_xlsx(blob[z:])
            if t:
                return {"loai": "excel", "nguon_stream": "zip-embed", **t}
        # ảnh (WMF/EMF/DIB) — không phải Compound File
        head = blob[:4].hex()
        if head == "d7cdc69a" or b" EMF" in blob[:64] or blob[:4] == b"\x01\x00\x00\x00":
            return {"loai": "anh", "ghi_chu": "WMF/EMF/DIB preview"}
        return {"loai": "khac", "ghi_chu": "khong magic CFB/zip"}
    if olefile is None:
        return {"loai": "khac", "ghi_chu": "thieu olefile"}
    try:
        ole = olefile.OleFileIO(io.BytesIO(blob[off:]))
    except Exception:
        return {"loai": "khac", "ghi_chu": "olefile mo loi"}
    try:
        streams = ["/".join(s) for s in ole.listdir()]
        pid = _progid(ole)
        # (2) BIFF .xls trực tiếp
        for wb in ("Workbook", "Book"):
            if ole.exists(wb):
                t = _rows_from_xls(ole.openstream(wb).read())
                if t:
                    return {"loai": "excel", "nguon_stream": wb, "progid": pid, **t}
        # (3) Package / CONTENTS / stream data -> file Excel LỒNG bên trong (bỏ header packager)
        for sname in streams:
            low = sname.lower()
            if low.endswith("compobj") or low.endswith("ole") or "olepres" in low or "summaryinformation" in low:
                continue
            try:
                data = ole.openstream(sname).read()
            except Exception:
                continue
            z = data.find(PKZIP)
            if z >= 0:
                t = _rows_from_xlsx(data[z:])
                if t:
                    return {"loai": "excel", "nguon_stream": sname + "/zip", "progid": pid, **t}
            c2 = data.find(CFBF)
            if c2 >= 0:
                try:
                    o2 = olefile.OleFileIO(io.BytesIO(data[c2:]))
                    for wb in ("Workbook", "Book"):
                        if o2.exists(wb):
                            t = _rows_from_xls(o2.openstream(wb).read())
                            if t:
                                o2.close()
                                return {"loai": "excel", "nguon_stream": sname + "/cfb", "progid": pid, **t}
                    o2.close()
                except Exception:
                    pass
        # có CFB nhưng không rút được bảng
        pl = pid.lower()
        if any(k in pl for k in ("picture", "bitmap", "dib", "brush", "metafile", "image")):
            return {"loai": "anh", "progid": pid}
        return {"loai": "khac", "progid": pid, "ghi_chu": "CFB khong co bang Excel"}
    finally:
        ole.close()


def doc_bang_ole(doc):
    """Duyệt MỌI OLE2FRAME (modelspace + paperspace), trả list dict CHỈ-ĐỌC.
    Mỗi phần tử: {handle, layer, khong_gian, loai, [sheet,nrows,ncols,rows,nguon_stream,progid]}.
    Fail-soft từng entity. KHÔNG áp to_unicode (caller làm, để module standalone-test được)."""
    out = []
    if doc is None:
        return out

    def _quet(space, ten_kg):
        try:
            it = list(space)
        except Exception:
            return
        for e in it:
            try:
                if e.dxftype() != "OLE2FRAME":
                    continue
            except Exception:
                continue
            rec = {"handle": None, "layer": None, "khong_gian": ten_kg}
            try:
                rec["handle"] = e.dxf.handle
                rec["layer"] = e.dxf.get("layer")
            except Exception:
                pass
            try:
                blob = e.binary_data()
            except Exception:
                blob = None
            try:
                rec.update(doc_tu_blob(blob))
            except Exception as ex:
                rec.update({"loai": "khac", "ghi_chu": "doc_tu_blob loi: %s" % ex})
            out.append(rec)

    try:
        _quet(doc.modelspace(), "modelspace")
    except Exception:
        pass
    try:
        for ln in doc.layout_names():
            if ln == "Model":
                continue
            try:
                _quet(doc.layout(ln), "paperspace")
            except Exception:
                continue
    except Exception:
        pass
    return out
